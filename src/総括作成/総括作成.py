import dataclasses
import glob
import os
import shutil

import openpyxl

from settings import ConstFullname
from src.common.const import Const総括作成用設定ファイルシート名, Constシート名
from src.common.openpyxl_util import is学習メモである

INT作業row = 3
INT作業column = 3
STR_ヘッダーセル範囲_目次 = "B2:H2"
STR_ヘッダーセル範囲_索引 = "B2:F2"
STR_セル範囲の一部_目次 = "B3:H"
STR_セル範囲の一部_索引 = "B3:F"
INPUT_FILE = ConstFullname.excel_input_file_総括作成用設定ファイル
OUTPUT_FILE = ConstFullname.excel_output_file_学習メモ総括


@dataclasses.dataclass
class WorkBookValuesObject:
    wb: str = dataclasses.field(default=None)
    ws: str = dataclasses.field(default=None)
    cell編集範囲: str = dataclasses.field(default=None)


def int_excel書き込み(ws総括メモ, tuple_cell, int作業row, xl_name):
    int作業row2 = int作業row
    for row in tuple_cell:
        int作業column = INT作業column
        ws総括メモ.cell(row=int作業row2, column=int作業column - 1).value = xl_name
        for col in row:
            ws総括メモ.cell(row=int作業row2, column=int作業column).value = col.value
            int作業column += 1
        int作業row2 += 1
    return int作業row2


def void_excelヘッダー作成(ws総括メモ, tuple_cell_学習メモ):
    cnt = 3
    ws総括メモ.cell(row=2, column=2).value = "excel保管場所"
    for row in tuple_cell_学習メモ:
        for col in row:
            ws総括メモ.cell(row=2, column=cnt).value = col.value
            cnt += 1


def void_学習メモ総括を作成():
    # インプットファイルがあるかどうかをチェック
    if not os.path.isfile(INPUT_FILE):
        print(f"インプットファイル[{INPUT_FILE}]が存在しません。")
        exit(1)

    # 設定ファイルからインプットフォルダ情報を取得
    wb = openpyxl.load_workbook(filename=INPUT_FILE, read_only=True)
    ws = wb[Const総括作成用設定ファイルシート名.str設定]
    while_itr = 3
    my_column = 2
    rg = ws.cell(column=my_column, row=while_itr)
    list対象フォルダ群 = []
    while rg.value is not None:
        if not os.path.isdir(rg.value):
            print(f"設定ファイル[{INPUT_FILE}]記載の対象のフォルダ[{rg.value}]が存在しません。")
            exit(1)
        list対象フォルダ群.append(rg.value)
        while_itr += 1
        rg = ws.cell(column=my_column, row=while_itr)

    if len(list対象フォルダ群) == 0:
        print(f"設定ファイル[{INPUT_FILE}]に対象フォルダの記載がありません。")
        exit(1)

    list_excelファイル群 = []
    for str_dir_name in list対象フォルダ群:
        files = glob.glob(f"{str_dir_name}\\*.xlsm")
        for file in files:
            list_excelファイル群.append(file)

    # 総括メモの作成とバックアップの作成
    if os.path.isfile(ConstFullname.excel_output_file_学習メモ総括):
        shutil.copy2(ConstFullname.excel_output_file_学習メモ総括, ConstFullname.excel_output_file_学習メモ総括bkp)

    # 学習メモであるものだけを抽出し、その内容を抽出してExcelにアウトプット
    wb総括メモ = openpyxl.Workbook()
    ws総括メモ_目次 = wb総括メモ.active
    ws総括メモ_目次.title = "目次"
    ws総括メモ_索引 = wb総括メモ.create_sheet(title="索引")
    int作業row_目次 = INT作業row
    int作業row_索引 = INT作業row
    isヘッダー未作成 = True
    for xl in list_excelファイル群:
        if is学習メモである(xl):
            # 学習メモをロード
            wb学習メモ = openpyxl.load_workbook(filename=xl, read_only=True)
            ws学習メモ_目次 = wb学習メモ[Constシート名.str_目次]
            ws学習メモ_索引 = wb学習メモ[Constシート名.str_索引]
            int_目次_max_row = ws学習メモ_目次.max_row
            int_索引_max_row = ws学習メモ_索引.max_row
            if int_目次_max_row <= 2 or int_索引_max_row <= 2: continue
            if isヘッダー未作成:
                void_excelヘッダー作成(ws総括メモ=ws総括メモ_目次, tuple_cell_学習メモ=ws学習メモ_目次[STR_ヘッダーセル範囲_目次])
                void_excelヘッダー作成(ws総括メモ=ws総括メモ_索引, tuple_cell_学習メモ=ws学習メモ_索引[STR_ヘッダーセル範囲_索引])
                isヘッダー未作成 = False

            int作業row_目次 = int_excel書き込み(
                ws総括メモ=ws総括メモ_目次, tuple_cell=ws学習メモ_目次[f"{STR_セル範囲の一部_目次}{int_目次_max_row}"],
                int作業row=int作業row_目次, xl_name=xl)
            int作業row_索引 = int_excel書き込み(
                ws総括メモ=ws総括メモ_索引, tuple_cell=ws学習メモ_索引[f"{STR_セル範囲の一部_索引}{int_索引_max_row}"],
                int作業row=int作業row_索引, xl_name=xl)

    wb総括メモ.save(ConstFullname.excel_output_file_学習メモ総括)


if __name__ == '__main__':
    void_学習メモ総括を作成()
